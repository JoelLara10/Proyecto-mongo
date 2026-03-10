# utils/backups.py
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError
import os
import json
import csv
import zipfile
import io
from datetime import datetime, date
from decimal import Decimal
from bson import ObjectId  # <--- AGREGAR ESTA LÍNEA
from bson.decimal128 import Decimal128
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
import bcrypt
from flask import current_app
from bd import get_db_connection
from datetime import datetime, date  # Ya debe estar

# Constantes
RUTA_CONFIG_AUTO = os.path.join('configuracion', 'copias', 'config_auto.json')
RUTA_CONTROL = os.path.join('configuracion', 'copias', 'control_backups.json')


def obtener_colecciones_mongo():
    """Obtiene todas las colecciones de MongoDB"""
    db = get_db_connection()
    return db.list_collection_names()


def list_backups():
    """Lista todos los archivos de backup"""
    carpeta = os.path.join(current_app.root_path, 'configuracion', 'copias')
    if not os.path.exists(carpeta):
        return []

    backups = []
    for file in os.listdir(carpeta):
        if file.startswith('backup_') and file.endswith(('.json', '.zip', '.xlsx', '.pdf')):
            filepath = os.path.join(carpeta, file)
            backups.append({
                'nombre': file,
                'fecha': datetime.fromtimestamp(os.path.getmtime(filepath)),
                'tamano': os.path.getsize(filepath)
            })
    return sorted(backups, key=lambda x: x['fecha'], reverse=True)


def cargar_control():
    ruta = os.path.join(current_app.root_path, RUTA_CONTROL)
    if not os.path.exists(ruta):
        return {}
    try:
        with open(ruta, 'r', encoding='utf-8') as f:
            return json.loads(f.read() or '{}')
    except:
        return {}


def guardar_control(data):
    ruta = os.path.join(current_app.root_path, RUTA_CONTROL)
    os.makedirs(os.path.dirname(ruta), exist_ok=True)
    with open(ruta, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4, default=str)


def cargar_config_auto():
    ruta = os.path.join(current_app.root_path, RUTA_CONFIG_AUTO)
    config_default = {
        'intervalo': 60,
        'tipo': 'completa',
        'formato': 'json',
        'auto_restore': False,
        'activo': True
    }
    if not os.path.exists(ruta):
        return config_default
    try:
        with open(ruta, 'r', encoding='utf-8') as f:
            return json.loads(f.read() or '{}')
    except:
        return config_default


def guardar_config_auto(config):
    ruta = os.path.join(current_app.root_path, RUTA_CONFIG_AUTO)
    os.makedirs(os.path.dirname(ruta), exist_ok=True)
    with open(ruta, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=4)


def validar_admin(username, password):
    """Valida credenciales de administrador"""
    db = get_db_connection()
    user = db['users'].find_one({"username": username})

    if not user:
        return False

    # Obtener la contraseña almacenada
    stored_password = user['password']

    # Si es string, convertir a bytes
    if isinstance(stored_password, str):
        stored_password = stored_password.encode('utf-8')

    # La contraseña ingresada siempre debe ser bytes
    password_bytes = password.encode('utf-8')

    # Verificar contraseña con bcrypt
    if not bcrypt.checkpw(password_bytes, stored_password):
        return False

    # Verificar rol de admin
    if user.get('role') != 'admin':
        return False

    return True


def limpiar_backups(max_por_tipo=4):
    carpeta = os.path.join(current_app.root_path, 'configuracion', 'copias')
    if not os.path.exists(carpeta):
        return

    backups = {}
    for f in os.listdir(carpeta):
        ruta = os.path.join(carpeta, f)
        if not os.path.isfile(ruta) or not f.startswith('backup_'):
            continue

        partes = f.split('_')
        if len(partes) >= 3:
            if partes[1] == 'auto':
                grupo = f"auto_{partes[2]}"
            elif partes[1] == 'manual':
                grupo = f"manual_{partes[2]}"
            else:
                grupo = f"manual_{partes[1]}"
            backups.setdefault(grupo, []).append(ruta)

    for grupo, archivos in backups.items():
        archivos.sort(key=os.path.getmtime, reverse=True)
        for archivo in archivos[max_por_tipo:]:
            try:
                os.remove(archivo)
            except Exception as e:
                current_app.logger.error(f"Error eliminando {archivo}: {str(e)}")


def realizar_backup(tipo, formato, colecciones_seleccionadas, es_auto=False):
    try:
        db = get_db_connection()
        carpeta = os.path.join(current_app.root_path, 'configuracion', 'copias')
        os.makedirs(carpeta, exist_ok=True)

        fecha = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        prefijo = 'auto' if es_auto else 'manual'
        nombre_archivo = f'backup_{prefijo}_{tipo}_{fecha}.{formato}'
        ruta_archivo = os.path.join(carpeta, nombre_archivo)

        # Filtrar por cambios si es diferencial/incremental
        if tipo in ['diferencial', 'incremental']:
            control = cargar_control()
            ultima_completa = control.get('ultima_completa')
            ultima_copia = control.get('ultima_copia')

            colecciones_con_cambios = []
            for coll in colecciones_seleccionadas:
                query = {}
                if tipo == 'diferencial' and ultima_completa:
                    query = {"updated_at": {"$gt": datetime.fromisoformat(ultima_completa)}}
                elif tipo == 'incremental' and ultima_copia:
                    query = {"updated_at": {"$gt": datetime.fromisoformat(ultima_copia)}}

                if query:
                    count = db[coll].count_documents(query)
                    if count > 0:
                        colecciones_con_cambios.append(coll)
                else:
                    colecciones_con_cambios.append(coll)
            colecciones_seleccionadas = colecciones_con_cambios

        if not colecciones_seleccionadas:
            return None

        # ============= FORMATO JSON =============
        if formato == 'json':
            respaldo = {}
            for coll in colecciones_seleccionadas:
                docs = list(db[coll].find())
                # Convertir ObjectId y otros tipos especiales a string
                for doc in docs:
                    doc['_id'] = str(doc['_id'])
                    # Convertir otros tipos especiales si es necesario
                    for key, value in doc.items():
                        if isinstance(value, ObjectId):
                            doc[key] = str(value)
                        elif isinstance(value, (datetime, date)):
                            doc[key] = value.isoformat()
                        elif isinstance(value, (Decimal, Decimal128)):
                            doc[key] = float(str(value))
                respaldo[coll] = docs

            with open(ruta_archivo, 'w', encoding='utf-8') as f:
                json.dump(respaldo, f, indent=4, ensure_ascii=False, default=str)

        # ============= FORMATO CSV (ZIP) =============
        elif formato == 'csv':
            try:
                with zipfile.ZipFile(ruta_archivo, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for coll in colecciones_seleccionadas:
                        docs = list(db[coll].find())
                        if docs:
                            # Preparar datos
                            all_fields = set()
                            for doc in docs:
                                for key in doc.keys():
                                    all_fields.add(key)

                            fieldnames = sorted(list(all_fields))

                            # Crear CSV en memoria
                            csv_buffer = io.StringIO()
                            writer = csv.DictWriter(csv_buffer, fieldnames=fieldnames, extrasaction='ignore')
                            writer.writeheader()

                            for doc in docs:
                                row = {}
                                for field in fieldnames:
                                    value = doc.get(field, '')
                                    if isinstance(value, ObjectId):
                                        value = str(value)
                                    elif isinstance(value, (datetime, date)):
                                        value = value.isoformat()
                                    elif isinstance(value, (Decimal, Decimal128)):
                                        value = float(str(value))
                                    elif value is None:
                                        value = ''
                                    else:
                                        value = str(value)
                                    row[field] = value
                                writer.writerow(row)

                            # Agregar al ZIP
                            zipf.writestr(f"{coll}.csv", csv_buffer.getvalue().encode('utf-8-sig'))

                current_app.logger.info(f"ZIP creado: {ruta_archivo}")

            except Exception as e:
                current_app.logger.error(f"Error en CSV: {str(e)}")
                raise e

        # ============= FORMATO EXCEL (CORREGIDO) =============
        elif formato == 'xlsx':
            try:
                wb = Workbook()
                # Eliminar hoja por defecto
                for sheet in wb.sheetnames:
                    wb.remove(wb[sheet])

                for coll in colecciones_seleccionadas:
                    docs = list(db[coll].find())
                    if docs:
                        # Crear nombre de hoja válido (Excel: máximo 31 chars, no []:*?/\)
                        sheet_name = coll[:31]
                        sheet_name = ''.join(c for c in sheet_name if c not in '[]:*?/\\')
                        if not sheet_name:
                            sheet_name = "Sheet"

                        ws = wb.create_sheet(title=sheet_name)

                        # Preparar datos
                        excel_docs = []
                        all_fields = set()

                        for doc in docs:
                            new_doc = {}
                            for key, value in doc.items():
                                clean_key = str(key).replace('.', '_').replace('$', '_')

                                if isinstance(value, ObjectId):
                                    new_doc[clean_key] = str(value)
                                elif isinstance(value, (datetime, date)):
                                    new_doc[clean_key] = value.strftime('%Y-%m-%d %H:%M:%S')
                                elif isinstance(value, (Decimal, Decimal128)):
                                    new_doc[clean_key] = float(str(value))
                                elif value is None:
                                    new_doc[clean_key] = ''
                                else:
                                    # Limpiar caracteres que Excel no acepta
                                    str_value = str(value)
                                    # Eliminar caracteres de control excepto tabs, newlines
                                    str_value = ''.join(
                                        char for char in str_value if ord(char) >= 32 or char in ['\n', '\r', '\t'])
                                    new_doc[clean_key] = str_value
                                all_fields.add(clean_key)
                            excel_docs.append(new_doc)

                        # Ordenar campos
                        fieldnames = sorted(list(all_fields))

                        # Escribir encabezados
                        ws.append(fieldnames)

                        # Escribir datos
                        for doc in excel_docs:
                            row = []
                            for field in fieldnames:
                                value = doc.get(field, '')
                                # Limitar longitud para evitar problemas
                                if isinstance(value, str) and len(value) > 32767:  # Límite de Excel
                                    value = value[:32767] + '...'
                                row.append(value)
                            ws.append(row)

                        # Ajustar ancho de columnas
                        for column in ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
                            ws.column_dimensions[column_letter].width = adjusted_width

                # Guardar Excel
                wb.save(ruta_archivo)

            except IllegalCharacterError as e:
                current_app.logger.error(f"Error de caracteres ilegales en Excel: {str(e)}")
                # Reintentar con limpieza más agresiva
                wb = Workbook()
                for sheet in wb.sheetnames:
                    wb.remove(wb[sheet])

                for coll in colecciones_seleccionadas:
                    docs = list(db[coll].find())
                    if docs:
                        sheet_name = coll[:31]
                        sheet_name = ''.join(c for c in sheet_name if c not in '[]:*?/\\')
                        if not sheet_name:
                            sheet_name = "Sheet"

                        ws = wb.create_sheet(title=sheet_name)

                        # Versión ultra-limpia para Excel
                        all_fields = set()
                        clean_docs = []

                        for doc in docs:
                            clean_doc = {}
                            for key, value in doc.items():
                                clean_key = str(key).replace('.', '_').replace('$', '_')
                                if isinstance(value, ObjectId):
                                    clean_doc[clean_key] = str(value)
                                elif isinstance(value, (datetime, date)):
                                    clean_doc[clean_key] = value.strftime('%Y-%m-%d %H:%M:%S')
                                elif isinstance(value, (Decimal, Decimal128)):
                                    clean_doc[clean_key] = float(str(value))
                                elif value is None:
                                    clean_doc[clean_key] = ''
                                else:
                                    # Limpieza extrema
                                    str_value = str(value)
                                    # Solo caracteres imprimibles ASCII
                                    str_value = ''.join(
                                        char for char in str_value if 32 <= ord(char) <= 126 or char in '\n\r\t')
                                    clean_doc[clean_key] = str_value
                                all_fields.add(clean_key)
                            clean_docs.append(clean_doc)

                        fieldnames = sorted(list(all_fields))
                        ws.append(fieldnames)

                        for doc in clean_docs:
                            row = [doc.get(field, '') for field in fieldnames]
                            ws.append(row)

                wb.save(ruta_archivo)

        # ============= FORMATO PDF =============
        elif formato == 'pdf':
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            from reportlab.lib.units import inch

            pdf_doc = SimpleDocTemplate(ruta_archivo, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []

            # Título
            title_style = styles['Title']
            story.append(Paragraph(f"Backup de Base de Datos - {fecha}", title_style))
            story.append(Spacer(1, 0.25 * inch))

            # Información del backup
            info_style = styles['Normal']
            story.append(Paragraph(f"Tipo: {tipo.upper()}", info_style))
            story.append(Paragraph(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", info_style))
            story.append(Spacer(1, 0.25 * inch))

            for coll in colecciones_seleccionadas:
                docs = list(db[coll].find().limit(50))

                # Título de colección
                story.append(Paragraph(f"Colección: {coll} ({len(docs)} documentos)", styles['Heading2']))
                story.append(Spacer(1, 0.1 * inch))

                if docs:
                    # Preparar datos para tabla
                    pdf_docs = []
                    all_fields = set()

                    for doc_item in docs:
                        new_doc = {}
                        for key, value in doc_item.items():
                            if isinstance(value, ObjectId):
                                new_doc[key] = str(value)[-8:]
                            elif isinstance(value, (datetime, date)):
                                new_doc[key] = value.strftime('%Y-%m-%d')
                            elif isinstance(value, (Decimal, Decimal128)):
                                new_doc[key] = f"{float(str(value)):.2f}"
                            elif isinstance(value, str) and len(value) > 30:
                                new_doc[key] = value[:27] + "..."
                            elif value is None:
                                new_doc[key] = ''
                            else:
                                new_doc[key] = str(value)
                            all_fields.add(key)
                        pdf_docs.append(new_doc)

                    fieldnames = sorted(list(all_fields))

                    if pdf_docs:
                        # Limitar a 8 columnas para PDF
                        if len(fieldnames) > 8:
                            display_fields = fieldnames[:8]
                        else:
                            display_fields = fieldnames

                        data = [display_fields]
                        for doc_item in pdf_docs:
                            row = [doc_item.get(field, '') for field in display_fields]
                            data.append(row)

                        col_widths = [(7.0 / len(display_fields)) * inch] * len(display_fields)

                        table = Table(data, colWidths=col_widths, repeatRows=1)
                        table.setStyle(TableStyle([
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                            ('FONTSIZE', (0, 0), (-1, -1), 7),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ]))

                        story.append(table)
                    else:
                        story.append(Paragraph("Sin datos", styles['Normal']))
                else:
                    story.append(Paragraph("Sin datos", styles['Normal']))

                story.append(Spacer(1, 0.2 * inch))

            pdf_doc.build(story)

        # Actualizar control
        control = cargar_control()
        ahora = datetime.now().isoformat()
        control['ultima_copia'] = ahora
        if tipo == 'completa':
            control['ultima_completa'] = ahora
        guardar_control(control)

        return nombre_archivo

    except Exception as e:
        current_app.logger.error(f"Error en backup: {str(e)}")
        import traceback
        traceback.print_exc()
        raise e

def restaurar_backup(nombre_archivo):
    try:
        carpeta = os.path.join(current_app.root_path, 'configuracion', 'copias')
        ruta_archivo = os.path.join(carpeta, nombre_archivo)

        if not os.path.exists(ruta_archivo):
            raise Exception("Archivo no encontrado")

        db = get_db_connection()
        ext = os.path.splitext(nombre_archivo)[1].lower()

        if ext == '.json':
            with open(ruta_archivo, 'r', encoding='utf-8') as f:
                data = json.load(f)
            for coll, docs in data.items():
                if docs:
                    db[coll].delete_many({})
                    db[coll].insert_many(docs)

        elif ext == '.zip':
            with zipfile.ZipFile(ruta_archivo, 'r') as zipf:
                for name in zipf.namelist():
                    if name.endswith('.csv'):
                        coll = name.replace('.csv', '')
                        with zipf.open(name) as csvfile:
                            reader = csv.DictReader(io.TextIOWrapper(csvfile, 'utf-8'))
                            docs = list(reader)
                            if docs:
                                db[coll].delete_many({})
                                db[coll].insert_many(docs)

        elif ext == '.xlsx':
            from openpyxl import load_workbook
            wb = load_workbook(ruta_archivo)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = [cell.value for cell in ws[1]]
                docs = []
                for row in ws.iter_rows(min_row=2):
                    doc = {}
                    for i, cell in enumerate(row):
                        if i < len(headers) and headers[i]:
                            doc[headers[i]] = cell.value
                    if doc:
                        docs.append(doc)
                if docs:
                    db[sheet_name].delete_many({})
                    db[sheet_name].insert_many(docs)

        return True

    except Exception as e:
        current_app.logger.error(f"Error restaurando: {str(e)}")
        raise e


def check_db_health():
    try:
        db = get_db_connection()
        db.command("ping")
        return True
    except Exception as e:
        current_app.logger.error(f"Error en health check: {str(e)}")
        return False


def job_backup_auto(app):
    with app.app_context():
        config = cargar_config_auto()
        if not config.get('activo', True):
            return

        if config.get('auto_restore', False) and not check_db_health():
            backups = list_backups()
            if backups:
                try:
                    restaurar_backup(backups[0]['nombre'])
                except Exception as e:
                    current_app.logger.error(f"Error en auto-restore: {str(e)}")

        try:
            colecciones = obtener_colecciones_mongo()
            nombre = realizar_backup(
                tipo=config['tipo'],
                formato=config['formato'],
                colecciones_seleccionadas=colecciones,
                es_auto=True
            )
            if nombre:
                limpiar_backups(max_por_tipo=4)
        except Exception as e:
            current_app.logger.error(f"Error en backup automático: {str(e)}")