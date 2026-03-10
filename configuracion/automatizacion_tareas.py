# configuracion/automatizacion_tareas.py
import os
import json
import csv
import zipfile
import io
from datetime import datetime, date
from decimal import Decimal

# MongoDB
from bson import ObjectId
from bson.decimal128 import Decimal128

# Excel
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import IllegalCharacterError

# PDF
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch

# Otros
import bcrypt
from flask import current_app
from bd import get_db_connection

# Constantes
RUTA_CONFIG_AUTO = os.path.join('configuracion', 'copias', 'config_auto.json')
RUTA_CONTROL = os.path.join('configuracion', 'copias', 'control_backups.json')


# ===============================
# FUNCIONES AUXILIARES
# ===============================

def obtener_colecciones_mongo():
    """Obtiene todas las colecciones de la base de datos MongoDB"""
    db = get_db_connection()
    return db.list_collection_names()


def list_backups():
    """Lista todos los archivos de backup en la carpeta configurada"""
    carpeta = os.path.join(current_app.root_path, 'configuracion', 'copias')
    if not os.path.exists(carpeta):
        return []

    backups = []
    for file in os.listdir(carpeta):
        if file.startswith('backup_') and file.endswith(('.json', '.zip', '.xlsx', '.pdf')):
            filepath = os.path.join(carpeta, file)
            backups.append({
                'nombre': file,
                'fecha': datetime.fromtimestamp(os.path.getmtime(filepath)),
                'tamano': os.path.getsize(filepath)
            })
    return sorted(backups, key=lambda x: x['fecha'], reverse=True)


def cargar_control():
    """Carga el archivo de control de backups"""
    ruta = os.path.join(current_app.root_path, RUTA_CONTROL)
    if not os.path.exists(ruta):
        return {}
    try:
        with open(ruta, 'r', encoding='utf-8') as f:
            data = f.read().strip()
            return json.loads(data) if data else {}
    except Exception as e:
        current_app.logger.error(f"Error cargando control: {str(e)}")
        return {}


def guardar_control(data):
    """Guarda el archivo de control de backups"""
    ruta = os.path.join(current_app.root_path, RUTA_CONTROL)
    os.makedirs(os.path.dirname(ruta), exist_ok=True)
    with open(ruta, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4, default=str)


def cargar_config_auto():
    """Carga la configuración de backup automático"""
    ruta = os.path.join(current_app.root_path, RUTA_CONFIG_AUTO)
    config_default = {
        'intervalo': 60,
        'tipo': 'completa',
        'formato': 'json',
        'auto_restore': False,
        'activo': True
    }

    if not os.path.exists(ruta):
        return config_default

    try:
        with open(ruta, 'r', encoding='utf-8') as f:
            data = f.read().strip()
            return json.loads(data) if data else config_default
    except Exception as e:
        current_app.logger.error(f"Error cargando config auto: {str(e)}")
        return config_default


def guardar_config_auto(config):
    """Guarda la configuración de backup automático"""
    ruta = os.path.join(current_app.root_path, RUTA_CONFIG_AUTO)
    os.makedirs(os.path.dirname(ruta), exist_ok=True)
    with open(ruta, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=4)


def validar_admin(username, password):
    """Valida credenciales de administrador"""
    db = get_db_connection()
    user = db['users'].find_one({"username": username})

    if not user:
        return False

    # Obtener la contraseña almacenada
    stored_password = user['password']

    # Si es string, convertir a bytes
    if isinstance(stored_password, str):
        stored_password = stored_password.encode('utf-8')

    # Verificar contraseña con bcrypt
    if not bcrypt.checkpw(password.encode('utf-8'), stored_password):
        return False

    # Verificar rol de admin
    if user.get('role') != 'admin':
        return False

    return True


def limpiar_backups(max_por_tipo=4):
    """Limpia backups antiguos manteniendo solo los más recientes por tipo"""
    carpeta = os.path.join(current_app.root_path, 'configuracion', 'copias')

    if not os.path.exists(carpeta):
        return

    backups = {}

    for f in os.listdir(carpeta):
        ruta = os.path.join(carpeta, f)
        if not os.path.isfile(ruta) or not f.startswith('backup_'):
            continue

        # Clasificar backup por tipo (manual/auto y completo/diferencial/incremental)
        partes = f.split('_')
        if len(partes) >= 3:
            if partes[1] == 'auto':
                grupo = f"auto_{partes[2]}"
            elif partes[1] == 'manual':
                grupo = f"manual_{partes[2]}"
            else:
                grupo = f"manual_{partes[1]}"

            backups.setdefault(grupo, []).append(ruta)

    # Eliminar backups excedentes
    for grupo, archivos in backups.items():
        archivos.sort(key=os.path.getmtime, reverse=True)
        for archivo in archivos[max_por_tipo:]:
            try:
                os.remove(archivo)
                current_app.logger.info(f"Backup eliminado: {os.path.basename(archivo)}")
            except Exception as e:
                current_app.logger.error(f"Error eliminando {archivo}: {str(e)}")


def check_db_health():
    """Verifica la salud de la base de datos"""
    try:
        db = get_db_connection()
        db.command("ping")
        return True
    except Exception as e:
        current_app.logger.error(f"Error en health check: {str(e)}")
        return False


# ===============================
# FUNCIÓN DE RESPALDO
# ===============================

def realizar_backup(tipo, formato, colecciones_seleccionadas, es_auto=False):
    """
    Realiza un backup de las colecciones seleccionadas

    Args:
        tipo: 'completa', 'diferencial', 'incremental'
        formato: 'json', 'csv', 'xlsx', 'pdf'
        colecciones_seleccionadas: lista de nombres de colecciones
        es_auto: True si es backup automático, False si es manual

    Returns:
        nombre_archivo: nombre del archivo generado o None si error
    """
    try:
        db = get_db_connection()
        carpeta = os.path.join(current_app.root_path, 'configuracion', 'copias')
        os.makedirs(carpeta, exist_ok=True)

        fecha = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        prefijo = 'auto' if es_auto else 'manual'

        # Determinar extensión según formato
        if formato == 'json':
            extension = 'json'
        elif formato == 'csv':
            extension = 'zip'
        elif formato == 'xlsx':
            extension = 'xlsx'
        elif formato == 'pdf':
            extension = 'pdf'
        else:
            extension = formato

        nombre_archivo = f'backup_{prefijo}_{tipo}_{fecha}.{extension}'
        ruta_archivo = os.path.join(carpeta, nombre_archivo)

        # Filtrar por cambios si es diferencial/incremental
        if tipo in ['diferencial', 'incremental']:
            control = cargar_control()
            ultima_completa = control.get('ultima_completa')
            ultima_copia = control.get('ultima_copia')

            colecciones_con_cambios = []
            for coll in colecciones_seleccionadas:
                query = {}
                if tipo == 'diferencial' and ultima_completa:
                    query = {"updated_at": {"$gt": datetime.fromisoformat(ultima_completa)}}
                elif tipo == 'incremental' and ultima_copia:
                    query = {"updated_at": {"$gt": datetime.fromisoformat(ultima_copia)}}

                if query:
                    count = db[coll].count_documents(query)
                    if count > 0:
                        colecciones_con_cambios.append(coll)
                else:
                    colecciones_con_cambios.append(coll)
            colecciones_seleccionadas = colecciones_con_cambios

        if not colecciones_seleccionadas:
            return None

        # ============= FORMATO JSON =============
        if formato == 'json':
            respaldo = {}
            for coll in colecciones_seleccionadas:
                docs = list(db[coll].find())
                # Convertir ObjectId y otros tipos especiales a string
                for doc in docs:
                    if '_id' in doc:
                        doc['_id'] = str(doc['_id'])
                    for key, value in doc.items():
                        if isinstance(value, ObjectId):
                            doc[key] = str(value)
                        elif isinstance(value, (datetime, date)):
                            doc[key] = value.isoformat()
                        elif isinstance(value, (Decimal, Decimal128)):
                            doc[key] = float(str(value))
                respaldo[coll] = docs

            with open(ruta_archivo, 'w', encoding='utf-8') as f:
                json.dump(respaldo, f, indent=4, ensure_ascii=False, default=str)

        # ============= FORMATO CSV (ZIP) =============
        elif formato == 'csv':
            with zipfile.ZipFile(ruta_archivo, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for coll in colecciones_seleccionadas:
                    docs = list(db[coll].find())
                    if docs:
                        # Preparar documentos para CSV
                        csv_docs = []
                        all_fields = set()

                        for doc in docs:
                            new_doc = {}
                            for key, value in doc.items():
                                clean_key = str(key).replace('.', '_').replace('$', '_')

                                if isinstance(value, ObjectId):
                                    new_doc[clean_key] = str(value)
                                elif isinstance(value, (datetime, date)):
                                    new_doc[clean_key] = value.isoformat()
                                elif isinstance(value, (Decimal, Decimal128)):
                                    new_doc[clean_key] = float(str(value))
                                elif value is None:
                                    new_doc[clean_key] = ''
                                else:
                                    str_value = str(value)
                                    str_value = ''.join(
                                        char for char in str_value if ord(char) >= 32 or char in ['\n', '\r', '\t'])
                                    new_doc[clean_key] = str_value
                                all_fields.add(clean_key)
                            csv_docs.append(new_doc)

                        fieldnames = sorted(list(all_fields))

                        # Escribir CSV
                        csv_buffer = io.StringIO()
                        writer = csv.DictWriter(csv_buffer, fieldnames=fieldnames, extrasaction='ignore')
                        writer.writeheader()

                        for doc in csv_docs:
                            row = {}
                            for field in fieldnames:
                                value = doc.get(field, '')
                                if isinstance(value, str):
                                    value = value.replace(',', ';').replace('\n', ' ').replace('\r', ' ')
                                row[field] = value
                            writer.writerow(row)

                        zipf.writestr(f"{coll}.csv", csv_buffer.getvalue().encode('utf-8-sig'))

        # ============= FORMATO EXCEL (XLSX) =============
        elif formato == 'xlsx':
            try:
                wb = Workbook()
                # Eliminar hoja por defecto
                for sheet in wb.sheetnames:
                    wb.remove(wb[sheet])

                for coll in colecciones_seleccionadas:
                    docs = list(db[coll].find())
                    if docs:
                        # Crear nombre de hoja válido
                        sheet_name = coll[:31]
                        invalid_chars = [':', '*', '?', '/', '\\', '[', ']']
                        for char in invalid_chars:
                            sheet_name = sheet_name.replace(char, '_')

                        if not sheet_name:
                            sheet_name = "Sheet1"

                        ws = wb.create_sheet(title=sheet_name)

                        # Preparar datos
                        excel_docs = []
                        all_fields = set()

                        for doc in docs:
                            new_doc = {}
                            for key, value in doc.items():
                                clean_key = str(key).replace('.', '_').replace('$', '_')

                                if isinstance(value, ObjectId):
                                    new_doc[clean_key] = str(value)
                                elif isinstance(value, (datetime, date)):
                                    new_doc[clean_key] = value.strftime('%Y-%m-%d %H:%M:%S')
                                elif isinstance(value, (Decimal, Decimal128)):
                                    new_doc[clean_key] = float(str(value))
                                elif value is None:
                                    new_doc[clean_key] = ''
                                else:
                                    str_value = str(value)
                                    str_value = ''.join(
                                        char for char in str_value if ord(char) >= 32 or char in ['\n', '\r', '\t'])
                                    if len(str_value) > 32767:
                                        str_value = str_value[:32767] + '...'
                                    new_doc[clean_key] = str_value
                                all_fields.add(clean_key)
                            excel_docs.append(new_doc)

                        fieldnames = sorted(list(all_fields))

                        # Escribir encabezados
                        ws.append(fieldnames)

                        # Escribir datos
                        for doc in excel_docs:
                            row = [doc.get(field, '') for field in fieldnames]
                            ws.append(row)

                wb.save(ruta_archivo)

            except Exception as e:
                current_app.logger.error(f"Error en Excel: {str(e)}")
                raise e

        # ============= FORMATO PDF =============
        elif formato == 'pdf':
            try:
                pdf_doc = SimpleDocTemplate(ruta_archivo, pagesize=letter)
                styles = getSampleStyleSheet()
                story = []

                # Título
                story.append(Paragraph(f"Backup de Base de Datos - {fecha}", styles['Title']))
                story.append(Spacer(1, 0.25 * inch))

                # Información del backup
                story.append(Paragraph(f"Tipo: {tipo.upper()}", styles['Normal']))
                story.append(Paragraph(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles['Normal']))
                story.append(Spacer(1, 0.25 * inch))

                for coll in colecciones_seleccionadas:
                    docs = list(db[coll].find().limit(50))

                    story.append(Paragraph(f"Colección: {coll} ({len(docs)} documentos)", styles['Heading2']))
                    story.append(Spacer(1, 0.1 * inch))

                    if docs:
                        # Preparar datos
                        pdf_docs = []
                        all_fields = set()

                        for doc_item in docs:
                            new_doc = {}
                            for key, value in doc_item.items():
                                clean_key = str(key).replace('.', '_').replace('$', '_')

                                if isinstance(value, ObjectId):
                                    new_doc[clean_key] = str(value)[-8:]
                                elif isinstance(value, (datetime, date)):
                                    new_doc[clean_key] = value.strftime('%Y-%m-%d')
                                elif isinstance(value, (Decimal, Decimal128)):
                                    new_doc[clean_key] = f"{float(str(value)):.2f}"
                                elif value is None:
                                    new_doc[clean_key] = ''
                                elif isinstance(value, str) and len(value) > 30:
                                    new_doc[clean_key] = value[:27] + "..."
                                else:
                                    new_doc[clean_key] = str(value)
                                all_fields.add(clean_key)
                            pdf_docs.append(new_doc)

                        fieldnames = sorted(list(all_fields))

                        if pdf_docs and fieldnames:
                            # Limitar a 8 columnas
                            if len(fieldnames) > 8:
                                display_fields = fieldnames[:8]
                            else:
                                display_fields = fieldnames

                            data = [display_fields]
                            for doc_item in pdf_docs:
                                row = [doc_item.get(field, '') for field in display_fields]
                                data.append(row)

                            col_widths = [(7.0 / len(display_fields)) * inch] * len(display_fields)

                            table = Table(data, colWidths=col_widths, repeatRows=1)
                            table.setStyle(TableStyle([
                                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                                ('FONTSIZE', (0, 0), (-1, -1), 7),
                                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ]))

                            story.append(table)
                        else:
                            story.append(Paragraph("Sin datos", styles['Normal']))
                    else:
                        story.append(Paragraph("Sin datos", styles['Normal']))

                    story.append(Spacer(1, 0.2 * inch))

                pdf_doc.build(story)

            except Exception as e:
                current_app.logger.error(f"Error en PDF: {str(e)}")
                raise e

        # Actualizar control
        control = cargar_control()
        ahora = datetime.now().isoformat()
        control['ultima_copia'] = ahora
        if tipo == 'completa':
            control['ultima_completa'] = ahora
        guardar_control(control)

        return nombre_archivo

    except Exception as e:
        current_app.logger.error(f"Error en backup: {str(e)}")
        import traceback
        traceback.print_exc()
        raise e


# ===============================
# FUNCIÓN DE RESTAURACIÓN (CORREGIDA)
# ===============================

def restaurar_backup(nombre_archivo):
    """
    Restaura un backup desde archivo

    Args:
        nombre_archivo: nombre del archivo a restaurar

    Returns:
        True si la restauración fue exitosa

    Raises:
        Exception: si hay error en la restauración
    """
    try:
        carpeta = os.path.join(current_app.root_path, 'configuracion', 'copias')
        ruta_archivo = os.path.join(carpeta, nombre_archivo)

        if not os.path.exists(ruta_archivo):
            raise Exception(f"Archivo no encontrado: {nombre_archivo}")

        db = get_db_connection()
        ext = os.path.splitext(nombre_archivo)[1].lower()

        current_app.logger.info(f"Iniciando restauración de {nombre_archivo} (formato: {ext})")

        # ============= RESTAURAR DESDE JSON =============
        if ext == '.json':
            with open(ruta_archivo, 'r', encoding='utf-8') as f:
                data = json.load(f)

            for coll, docs in data.items():
                if docs:
                    current_app.logger.info(f"Restaurando colección {coll} con {len(docs)} documentos")

                    # Limpiar colección
                    db[coll].delete_many({})

                    # Preparar documentos
                    for doc in docs:
                        # Restaurar ObjectId
                        if '_id' in doc and isinstance(doc['_id'], str):
                            try:
                                doc['_id'] = ObjectId(doc['_id'])
                            except:
                                pass

                        # Restaurar fechas
                        for key, value in doc.items():
                            if isinstance(value, str):
                                # Intentar convertir fechas ISO a datetime
                                try:
                                    if 'T' in value and len(value) > 10:
                                        doc[key] = datetime.fromisoformat(value.replace('Z', '+00:00'))
                                except:
                                    pass

                    # Insertar documentos
                    db[coll].insert_many(docs)

            current_app.logger.info(f"Restauración desde JSON completada: {len(data)} colecciones")

        # ============= RESTAURAR DESDE ZIP (CSV) =============
        elif ext == '.zip':
            with zipfile.ZipFile(ruta_archivo, 'r') as zipf:
                for name in zipf.namelist():
                    if name.endswith('.csv'):
                        coll = name.replace('.csv', '')
                        current_app.logger.info(f"Restaurando colección {coll} desde {name}")

                        with zipf.open(name) as csvfile:
                            # Leer CSV
                            content = csvfile.read().decode('utf-8-sig')
                            reader = csv.DictReader(io.StringIO(content))
                            docs = list(reader)

                            if docs:
                                # Limpiar colección
                                db[coll].delete_many({})

                                # Convertir tipos
                                for doc in docs:
                                    for key, value in doc.items():
                                        # Intentar convertir números
                                        if value and isinstance(value, str):
                                            try:
                                                if '.' in value:
                                                    doc[key] = float(value)
                                                else:
                                                    doc[key] = int(value)
                                            except:
                                                pass

                                # Insertar documentos
                                db[coll].insert_many(docs)
                                current_app.logger.info(f"  → {len(docs)} documentos restaurados")

        # ============= RESTAURAR DESDE EXCEL =============
        elif ext == '.xlsx':
            from openpyxl import load_workbook

            wb = load_workbook(ruta_archivo, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                current_app.logger.info(f"Restaurando colección {sheet_name} desde Excel")

                # Obtener encabezados
                headers = []
                for cell in ws[1]:
                    if cell.value:
                        headers.append(str(cell.value))

                if not headers:
                    current_app.logger.warning(f"  → No hay encabezados en la hoja {sheet_name}")
                    continue

                # Leer datos
                docs = []
                for row in ws.iter_rows(min_row=2):
                    doc = {}
                    has_data = False

                    for i, cell in enumerate(row):
                        if i < len(headers) and cell.value is not None:
                            value = cell.value
                            # Convertir tipos
                            if isinstance(value, (datetime, date)):
                                value = value.isoformat()
                            doc[headers[i]] = value
                            has_data = True

                    if has_data:
                        docs.append(doc)

                if docs:
                    # Limpiar colección
                    db[sheet_name].delete_many({})

                    # Insertar documentos
                    db[sheet_name].insert_many(docs)
                    current_app.logger.info(f"  → {len(docs)} documentos restaurados")

        # ============= PDF NO SE PUEDE RESTAURAR =============
        elif ext == '.pdf':
            raise Exception("Los archivos PDF son solo para visualización, no contienen datos recuperables")

        else:
            raise Exception(f"Formato no soportado para restauración: {ext}")

        current_app.logger.info(f"Restauración de {nombre_archivo} completada exitosamente")
        return True

    except Exception as e:
        current_app.logger.error(f"Error restaurando {nombre_archivo}: {str(e)}")
        import traceback
        traceback.print_exc()
        raise e


# ===============================
# JOB AUTOMÁTICO
# ===============================

def job_backup_auto(app):
    """Job programado para backups automáticos"""
    with app.app_context():
        config = cargar_config_auto()

        if not config.get('activo', True):
            return

        # Verificar auto-restore si está activado
        if config.get('auto_restore', False):
            if not check_db_health():
                backups = list_backups()
                if backups:
                    latest = backups[0]['nombre']
                    current_app.logger.info(f"Restaurando desde {latest} por error en BD")
                    try:
                        restaurar_backup(latest)
                        current_app.logger.info("Restauración automática completada")
                    except Exception as e:
                        current_app.logger.error(f"Error en auto-restore: {str(e)}")

        # Realizar backup automático
        try:
            colecciones = obtener_colecciones_mongo()
            nombre = realizar_backup(
                tipo=config['tipo'],
                formato=config['formato'],
                colecciones_seleccionadas=colecciones,
                es_auto=True
            )
            if nombre:
                current_app.logger.info(f"Backup automático creado: {nombre}")
                limpiar_backups(max_por_tipo=4)
        except Exception as e:
            current_app.logger.error(f"Error en backup automático: {str(e)}")